#!/usr/bin/env python3
"""
Export the /compare table to output/Sealshot-Comparison.xlsx for editing.

Round-trips: edit the spreadsheet, then run output/import-comparison-xlsx.py to
write it back into src/data/comparison.json, which is what the page renders from.
A one-way export would be a dead end — the point is to edit in Excel.

    python3 output/export-comparison-xlsx.py

Layout of the Comparison sheet, which the importer relies on:

  row 1            header — Capability, Category, then one column per app
  feature rows     label in column A, its section in column B, one value per app
  footnote markers a trailing " [n]" inside a cell, e.g. "Opaque solid fill [1]"

Sections come from the Category column and keep first-seen order, so reordering
rows reorders the page. Add or rename rows and categories freely. To add an app
you need to touch the JSON, because a column also carries a key and its source
links — the importer refuses an unknown column rather than guessing.
"""
import json
import pathlib
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA = ROOT / 'src' / 'data' / 'comparison.json'
OUT = pathlib.Path(__file__).resolve().parent / 'Sealshot-Comparison.xlsx'

INK = '0B1220'
ACCENT = 'C2410C'
BAND = 'F4FBF8'
TINT = 'E8EFEC'
FAINT = 'DBE3EE'

d = json.loads(DATA.read_text())
apps = d['apps']

wb = Workbook()

# ── Comparison ───────────────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Comparison'

thin = Side(style='thin', color=FAINT)
edge = Border(bottom=thin)
wrap = Alignment(vertical='top', wrap_text=True)

ws.append(['Capability', 'Category'] + [a['name'] for a in apps])
for i in range(1, len(apps) + 3):
    c = ws.cell(row=1, column=i)
    c.font = Font(bold=True, size=11, color=ACCENT if i == len(apps) + 1 else INK)
    c.border = Border(bottom=Side(style='medium', color=INK))
    c.alignment = wrap

row = 2
for group in d['groups']:
    for r in group['rows']:
        ws.cell(row=row, column=1, value=r['label']).font = Font(bold=True, size=10)
        ws.cell(row=row, column=2, value=group['title'])
        for i, a in enumerate(apps, start=3):
            note = r.get(a['key'] + 'Note')
            text = r.get(a['key'], '')
            ws.cell(row=row, column=i, value=f'{text} [{note}]' if note else text)
        for i in range(1, len(apps) + 3):
            c = ws.cell(row=row, column=i)
            c.alignment = wrap
            c.border = edge
            if i == len(apps) + 2:                    # our column, tinted as on the page
                c.fill = PatternFill('solid', fgColor=BAND)
        row += 1

ws.freeze_panes = 'C2'
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 14
for i in range(3, len(apps) + 3):
    ws.column_dimensions[get_column_letter(i)].width = 34

# ── Notes ────────────────────────────────────────────────────────────────────
nws = wb.create_sheet('Notes')
nws.append(['#', 'Footnote'])
for c in nws[1]:
    c.font = Font(bold=True)
for i, note in enumerate(d['notes'], start=1):
    nws.append([i, re.sub(r'\s+', ' ', note).strip()])
nws.column_dimensions['A'].width = 5
nws.column_dimensions['B'].width = 110
for r in nws.iter_rows(min_row=2):
    r[1].alignment = wrap

# ── Meta ─────────────────────────────────────────────────────────────────────
mws = wb.create_sheet('Meta')
mws.append(['Field', 'Value'])
for c in mws[1]:
    c.font = Font(bold=True)
mws.append(['checkedOn', d['checkedOn']])
mws.append(['', ''])
mws.append(['Legend term', 'Meaning'])
mws['A4'].font = Font(bold=True)
mws['B4'].font = Font(bold=True)
for term, meaning in d['legend']:
    mws.append([term, meaning])
mws.append(['', ''])
mws.append(['App', 'Sources'])
for a in apps:
    mws.append([a['name'], ' · '.join(f'{label}: {url}' for label, url in a['sources'])])
mws.column_dimensions['A'].width = 20
mws.column_dimensions['B'].width = 110

wb.save(OUT)
rows = sum(len(g['rows']) for g in d['groups'])
print(f'wrote {OUT.relative_to(ROOT)}')
print(f'  {rows} feature rows in {len(d["groups"])} sections, {len(apps)} apps, '
      f'{len(d["notes"])} footnotes')
print('  edit it, then: python3 output/import-comparison-xlsx.py')
