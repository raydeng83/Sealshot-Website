#!/usr/bin/env python3
"""
Read output/Sealshot-Comparison.xlsx back into src/data/comparison.json.

The page renders from the JSON, so this is what makes editing in Excel actually
mean something. Rebuild the site afterwards to see it.

    python3 output/import-comparison-xlsx.py
    python3 output/import-comparison-xlsx.py --dry-run          # show the diff only
    python3 output/import-comparison-xlsx.py --file other.xlsx  # any workbook

Two layouts are accepted, because both are reasonable to edit by hand:

  Capability | Category | <apps…>   grouping comes from the Category column
  Feature    | <apps…>              grouping comes from section rows — a title in
                                    column A with every app column empty

The header row is found by scanning for a first cell of "Capability" or "Feature",
so leading title and subtitle rows are fine. Blank rows are skipped, so clearing a
row in Excel removes it.

Refuses rather than guesses:
  - unknown app columns, or a missing one
  - a footnote marker [n] with no matching row on the Notes sheet
  - a feature row appearing before any section heading
  - an empty table, which would silently gut the page
App keys and source links are not in the spreadsheet, so they are preserved from
the existing JSON and matched by column name.
"""
import json
import pathlib
import re
import sys

from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA = ROOT / 'src' / 'data' / 'comparison.json'
XLSX = pathlib.Path(__file__).resolve().parent / 'Sealshot-Comparison.xlsx'

NOTE_RE = re.compile(r'\s*\[(\d+)\]\s*$')


def fail(msg: str) -> None:
    print(f'error: {msg}', file=sys.stderr)
    raise SystemExit(1)


def main() -> int:
    dry = '--dry-run' in sys.argv
    src = XLSX
    if '--file' in sys.argv:
        src = pathlib.Path(sys.argv[sys.argv.index('--file') + 1]).expanduser()
    if not src.exists():
        fail(f'{src} not found — run export-comparison-xlsx.py first')

    existing = json.loads(DATA.read_text())
    by_name = {a['name']: a for a in existing['apps']}

    wb = load_workbook(src, data_only=True)
    ws = wb['Comparison'] if 'Comparison' in wb.sheetnames else wb[wb.sheetnames[0]]

    # Find the header row rather than assuming row 1: a working spreadsheet
    # usually has a title and a note above the table.
    header, header_row = None, None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        first = str(row[0]).strip() if row and row[0] is not None else ''
        if first in ('Capability', 'Feature'):
            header, header_row = list(row), i
            break
    if header is None:
        fail('no header row found — column A of it must read "Capability" or "Feature"')

    has_category = len(header) > 1 and str(header[1] or '').strip() == 'Category'
    first_app_col = 2 if has_category else 1
    names = [(h or '').strip() for h in header[first_app_col:] if (h or '').strip()]
    unknown = [n for n in names if n not in by_name]
    if unknown:
        fail(f'unknown app column(s): {unknown}. Add the app to comparison.json '
             f'first — a column also needs a key and its source links.')
    missing = [n for n in by_name if n not in names]
    if missing:
        fail(f'column(s) removed from the sheet: {missing}. Remove them from '
             f'comparison.json too if that was deliberate.')

    # Notes sheet first, so footnote references can be validated as we read cells.
    notes = []
    if 'Notes' in wb.sheetnames:
        for row in wb['Notes'].iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None and str(row[1] or '').strip():
                notes.append(str(row[1]).strip())

    groups, current, by_title = [], None, {}
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        label = str(r[0]).strip() if r and r[0] is not None else ''
        values = [str(v).strip() if v is not None else ''
                  for v in r[first_app_col:first_app_col + len(names)]]
        if not label and not any(values):
            continue
        if has_category:
            title = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ''
            if not title:
                fail(f'row "{label}" has no Category')
            if title not in by_title:
                by_title[title] = {'title': title, 'rows': []}
                groups.append(by_title[title])       # categories keep first-seen order
            current = by_title[title]
        else:
            if label and not any(values):                   # section heading
                current = {'title': label, 'rows': []}
                groups.append(current)
                continue
            if current is None:
                fail(f'feature row "{label}" appears before any section heading')
        entry: dict[str, object] = {'label': label}
        for name, value in zip(names, values):
            key = by_name[name]['key']
            m = NOTE_RE.search(value)
            note = None
            if m:
                note = int(m.group(1))
                if note < 1 or note > len(notes):
                    fail(f'"{label}" / {name} cites footnote [{note}], but the Notes '
                         f'sheet has {len(notes)} row(s)')
                value = NOTE_RE.sub('', value)
            # value before its note, so the JSON key order stays readable and the
            # file is stable across repeated imports
            entry[key] = value
            if note is not None:
                entry[key + 'Note'] = note
        current['rows'].append(entry)

    total = sum(len(g['rows']) for g in groups)
    if not total:
        fail('no feature rows found; refusing to write an empty table')

    out = dict(existing)
    out['groups'] = groups
    out['notes'] = notes or existing['notes']

    if 'Meta' in wb.sheetnames:
        mws = wb['Meta']
        for row in mws.iter_rows(min_row=2, max_row=3, values_only=True):
            if row and str(row[0] or '').strip() == 'checkedOn' and row[1]:
                out['checkedOn'] = str(row[1]).strip()

    # Report what moved before touching anything.
    old_cells = {(g['title'], r['label'], k): v
                 for g in existing['groups'] for r in g['rows']
                 for k, v in r.items() if k != 'label'}
    new_cells = {(g['title'], r['label'], k): v
                 for g in groups for r in g['rows']
                 for k, v in r.items() if k != 'label'}
    # only cells present in both — otherwise a new row is reported twice,
    # once as changed and once as new, and the counts stop meaning anything
    changed = [k for k in new_cells if k in old_cells and old_cells[k] != new_cells[k]]
    added = [k for k in new_cells if k not in old_cells]
    removed = [k for k in old_cells if k not in new_cells]

    print(f'  {total} feature rows in {len(groups)} sections, {len(notes)} footnotes')
    print(f'  changed: {len(changed)}   new: {len(added)}   gone: {len(removed)}')
    for g, label, key in changed[:12]:
        if (g, label, key) in old_cells:
            print(f'    {label} / {key}: {old_cells[(g, label, key)]!r} -> '
                  f'{new_cells[(g, label, key)]!r}')
    if len(changed) > 12:
        print(f'    … and {len(changed) - 12} more')
    referenced = {v for g in groups for r in g['rows']
                  for k, v in r.items() if k.endswith('Note')}
    orphans = [i for i in range(1, len(out['notes']) + 1) if i not in referenced]
    if orphans:
        print(f'  warning: footnote(s) {orphans} are no longer cited by any cell. '
              f'They will still render. Re-add a reference or delete them.')

    if out['checkedOn'] != existing['checkedOn']:
        print(f'  checkedOn: {existing["checkedOn"]!r} -> {out["checkedOn"]!r}')
    elif changed or added or removed:
        print('  note: cells changed but checkedOn did not. Update it on the Meta '
              'sheet so the page does not claim a stale verification date.')

    if dry:
        print('\n--dry-run: nothing written.')
        return 0

    DATA.write_text(json.dumps(out, indent=2, ensure_ascii=False) + '\n')
    print(f'\nwrote {DATA.relative_to(ROOT)} — rebuild the site to see it')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
